from openpyxl import  Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Color, Alignment, Border, Side, GradientFill, PatternFill
from openpyxl.cell import MergedCell

import datetime
import os
from date_utils import get_full_month_th

wb = Workbook()
ws = wb.active
ws.title = 'Staff'



# Header
current_month = datetime.datetime.now()

# ws['A1'] = f"รายชื่อพนักงานใหม่ประจำเดือน {current_month}"
# ws['A1'].font = Font(name='Tahoma', b=True, size=18)

ws.cell(row=1, column=1).value = f"รายชื่อพนักงานใหม่ประจำเดือน {get_full_month_th(current_month.month)} {current_month.year}"
ws.cell(row=1, column=1).font = Font(name='Tahoma', b=True, size=18)

# Merge cell
end_header_column_letter = get_column_letter(6)
# ws.merge_cells(f'A1:F1')
ws.merge_cells(f'A1:{end_header_column_letter}1')

ws.cell(row=2, column=1).value = f"AA"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

merged_cell = ws.cell(row=1, column=1)
normal_cell = ws.cell(row=2, column=1)
# ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)


# Dimension
ws.column_dimensions['A'].width = 25.5 # set column width
ws.row_dimensions[1].height = 50  # set row height

# Alignment
ws.cell(row=1, column=1).alignment = Alignment(
    horizontal='center',
    vertical='center',
  )

# Number format
ws.cell(row=3, column=1).value = 10_000_000
ws.cell(row=3, column=1).number_format = 'QQQ#,##0.00' # format from excel

# Sub header
ws.cell(row=4, column=1).value = 'บริษัท ABC จำกัด'
ws.cell(row=4, column=1).font = Font(i=True , color='00FF0000')

ws.cell(row=4, column=6).value = 'ทั้งหมด 3 คน'

border_1 = Border(left=Side(border_style='thin',color='FF000000'),
                 right=Side(border_style='thin', color='FF000000'),
                 top=Side(border_style='thin', color='FF000000'),
                 bottom=Side(border_style='thin', color='FF000000'),
                )
bold_font = Font(b=True)

# Fields
fields = ["รหัสพนักงาน",	"ตำแหน่ง",	"คำนำหน้าชื่อ",	"ชื่อ",	"นามสกุล",	"แผนก"]
for at_column, field in enumerate(fields, start=1):
    ws.cell(row=5, column=at_column).value = field
    ws.cell(row=5, column=at_column).font = bold_font
    ws.cell(row=5, column=at_column).border = border_1
    ws.cell(row=5, column=at_column).fill = PatternFill("solid", fgColor="DDDDDD" )


# Data
data = [
    ['A0001',	'Manager',	'นาย',	'สมชาย',	'ดี',	'Development'],
    ['A0002',	'Manager',	'นาย',	'สมหมาย',	'ดีมาก',	'Development'],
    ['A0003',	'Senior Developer',	'น.ส.',	'เอ',	'เอบีซีดี',	'Development'],
]
for at_row, row in enumerate(data, start=6):
    for at_column, value in enumerate(row, start=1):
      ws.cell(row=at_row, column=at_column).value = value


ws.cell(row=10, column=3).value = 1
ws.cell(row=11, column=3).value = 11
ws.cell(row=12, column=3).value = 15
ws.cell(row=13, column=3).value = 19
ws.cell(row=14, column=3).value = 17
ws.cell(row=15, column=3).value = "=SUM(C10:C14)"

# Inserting
# ws.insert_rows(5, amount=6)
# ws.cell(row=5, column=1).value = 6666
# ws.cell(row=6, column=1).value = 11

# ws.insert_cols(3, amount=4)

# delete
# ws.delete_cols(3, amount=3)
# ws.delete_rows(1, amount=5)

# Save
file = r"staff_result2.xlsx"
if os.path.exists(file):
    os.remove(file)
wb.save(file)
wb.close()



wb1 = load_workbook("aaa.xlsx", data_only=True)
ws = wb1[wb1.sheetnames[0]]
print(ws['A8'].value)
wb.close()