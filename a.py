import json
import os
from openpyxl import load_workbook


def readrow(row: list):
    # return [
    #     '' if cv.value is None else 
    #     cv.value.strip() if type(cv.value) is str else 
    #     cv.value
    #     for cv in row
    # ]
  return [ cell_value(cell) for cell in row ]


def cell_value(cell: object):
    
    if not (cv:=cell.value):
        return ''
    elif type(cv) is str:
        return cv.strip()
    else:
        return cv


file = r"D:\Projects\Road to Advanced Python\Chapter 2\data\excel1.xlsx"
file = os.path.join(os.path.dirname(__file__), 'data', 'excel1.xlsx')

print("__file__", __file__)
wb = load_workbook(filename=file)
print(wb)

ws = wb['Levels']


map_levels = {}
for row in ws.iter_rows(min_row=2):
    row = readrow(row)
    lv = row[0]
    name = row[1]
    en_name = row[2]
    if not row[0]:
        continue
    # if not name or not en_name:
    #     continue
    print(row)
    
    # if name not in map_levels:
    #     map_levels[name] = [en_name, lv]

# print(map_levels)
# print(json.dumps(map_levels, indent=4, ensure_ascii=False))

wb.close()