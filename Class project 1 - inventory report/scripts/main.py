from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import os
from excel_utils import readrow, check_file_exist, get_header_by_row, overwrite_excel, map_header_value, delete_sheet
from typing import Dict
from config import Path, MasterField, TransactionOutField
import datetime
from date_utils import TH_MONTHS, get_full_month_th_with_year


class Master:

    def __init__(self, product_code: str, product_name: str, unit: str, price_per_unit: float | int, inventory_level: float | int) -> None:
        self.product_code = product_code
        self.product_name = product_name
        self.unit = unit
        self.price_per_unit = price_per_unit
        self.inventory_level = inventory_level



class TransactionOut:

    # product_code = 'product_code'
    # product_name = 'product_name'

    def __init__(self, product_code: str, product_name: str, unit: str, price_per_unit: float, qty: int, vat: float, excluding_vat: float, including_vat: float) -> None:
        self.product_code = (product_code, TransactionOutField.PRODUCT_CODE)
        self.product_name = (product_name, TransactionOutField.PRODUCT_NAME)
        self.unit = (unit, TransactionOutField.UNIT)
        self.price_per_unit = [price_per_unit, TransactionOutField.PRICE_PER_UNIT]
        self.qty = [qty, TransactionOutField.QTY]
        self.vat = [vat, TransactionOutField.VAT]
        self.excluding_vat = [excluding_vat, TransactionOutField.EXCLUDING_VAT]
        self.including_vat = [including_vat, TransactionOutField.INCLUDING_VAT]

    def __repr__(self) -> str:
        return f"code:{self.product_code} >>>> qty:{self.qty}"

    def __str__(self) -> str:
        return self.__repr__()

class TransactionIn:

    def __init__(self, product_code: str, product_name: str, unit: str, qty: int) -> None:
        self.product_code = (product_code, TransactionOutField.PRODUCT_CODE)
        self.product_name = (product_name, TransactionOutField.PRODUCT_NAME)
        self.unit = (unit, TransactionOutField.UNIT)
        self.qty = [qty, TransactionOutField.QTY]

    def __repr__(self) -> str:
        return f"code:{self.product_code} >>>> qty:{self.qty}"

    def __str__(self) -> str:
        return self.__repr__()


class YearlyReport:

    def __init__(self, save_file:str) -> None:
        self.save_file = save_file
        self.map_summary = {
            TransactionOutField.VAT: 0,
            TransactionOutField.EXCLUDING_VAT: 0,
            TransactionOutField.INCLUDING_VAT: 0,
        }
        self.headers = TransactionOutField.MAP_HEADERS
        self.last_row = 1


    def initial_workbook(self):
        self.wb = Workbook()

    def set_constant_cells(self, date: datetime.datetime):

        sheetname = date.strftime('%b %Y')
        if sheetname not in self.wb.sheetnames:
            self.ws = self.wb.create_sheet(title=sheetname)
        else:
            self.ws = self.wb.active
            self.ws.title = sheetname


        # set header
        self.ws['A2'] = f'รายงานสรุปรายการขายสินค้า ประจำเดือน {get_full_month_th_with_year(month=date.month, year=date.year)}'

        # set fields
        for at_column, header in enumerate(self.headers, start=1):
            self.ws.cell(row=4, column=at_column).value = header
        
        self.last_row = 4

    def set_data(self, map_master, map_transaction_out):
        sorted_product_code = sorted(list(map_master.keys()))
        # set data
        for at_row, product_code in enumerate(sorted_product_code, start=5):
            # get TransactionOut
            self.last_row += 1
            trx = map_transaction_out[product_code]
            for at_column, header_mapper in enumerate(self.headers.values(), start=1):
                value = map_header_value(object_attributes=trx.__dict__, header_mapper=header_mapper)
                self.ws.cell(row=at_row, column=at_column).value = value

                # increase value
                if header_mapper in self.map_summary:
                    self.map_summary[header_mapper] += value
    
    def set_summary(self):
        # set summary
        summary_row = self.last_row + 2
        
        self.ws.cell(row=summary_row, column=2).value = 'รวม'
        self.ws.cell(row=summary_row, column=5).value = f'=SUM(E5:E{self.last_row})'
        self.ws.cell(row=summary_row, column=6).value = f'=SUM(F5:F{self.last_row})'
        self.ws.cell(row=summary_row, column=7).value = f'=SUM(G5:G{self.last_row})'

    def remove_sheet(self):
        # sheet = 'Sheet'
        # if sheet in self.wb.sheetnames:
        #     self.wb.remove_sheet(sheet)
        delete_sheet(wb=self.wb, sheet='Sheet')

    def save(self):
        overwrite_excel(wb=self.wb, save_file=self.save_file )
        self.wb.close()



def read_master(file_master: str = Path.MASTER_FILE):
    """
    read master excel in Path.MASTER_FILE

    params: 
      file_master: str = absolute path file 

    return: map_master: Dict[str, Master]
    """

    map_master: Dict[str, Master] = {}

    # check file exist

    check_file_exist(file=file_master)

    # if not os.path.exists(file_master):
    #   raise Exception(f'Not found file {file_master}')

    # open workbook
    wb = load_workbook(filename=file_master, data_only=True)
    ws = wb[MasterField.MASTER_SHEET]

    # read header
    header = get_header_by_row(ws=ws, row_index=1)
    # header = []
    # for row in ws.iter_rows(min_row=1, max_row=1):
    #     header = readrow(row)

    # print(header)

    for row in ws.iter_rows(min_row=2):
        row = readrow(row)

        product_code = row[header.index(MasterField.PRODUCT_CODE)]
        master = Master(
            product_code=row[header.index(MasterField.PRODUCT_CODE)],
            product_name=row[header.index(MasterField.PRODUCT_NAME)],
            unit=row[header.index(MasterField.UNIT)],
            price_per_unit=row[header.index('Price/unit')],  # hard code
            inventory_level=row[4]
        )
        if product_code not in map_master:
            map_master[product_code] = master

    wb.close()

    return map_master

def read_transaction_out(map_master: dict, date: datetime.datetime, input_folder: str = Path.FOLDER_INPUT):
    """
    read transaction excel in Path.FOLDER_INPUT

    params: 
        map_master: Dict[str, Master] 
        input_folder: str = path of folder

    return: map_transaction_out: Dict[str, TransactionOut]
    """

    map_transaction_out: Dict[str, TransactionOut] = {}

    file_trx_out = os.path.join(input_folder, f"transaction out - {date.strftime('%Y %m')}.xlsx")

    check_file_exist(file=file_trx_out)

    # open workbook
    wb = load_workbook(filename=file_trx_out)
    ws = wb[wb.sheetnames[0]]

    # read header
    header = get_header_by_row(ws=ws, row_index=1)

    for row in ws.iter_rows(min_row=2):
        row = readrow(row)

        product_code = row[header.index(TransactionOutField.PRODUCT_CODE)]
        product_name = row[header.index(TransactionOutField.PRODUCT_NAME)]
        unit = row[header.index(TransactionOutField.UNIT)]
        price_per_unit = row[header.index(TransactionOutField.PRICE_PER_UNIT)]
        qty = row[header.index(TransactionOutField.QTY)]
        vat = row[header.index(TransactionOutField.VAT)]
        excluding_vat = row[header.index(TransactionOutField.EXCLUDING_VAT)]
        including_vat = row[header.index(TransactionOutField.INCLUDING_VAT)]

        if product_code not in map_master:
            continue

        if product_code not in map_transaction_out:
            map_transaction_out[product_code] = TransactionOut(
                product_code=product_code,
                product_name=product_name,
                unit=unit,
                price_per_unit=price_per_unit,
                qty=qty,
                vat=vat,
                excluding_vat=excluding_vat,
                including_vat=including_vat
            )

        else:
            map_transaction_out[product_code].qty[0] += qty
            map_transaction_out[product_code].vat[0] += vat
            map_transaction_out[product_code].excluding_vat[0] += excluding_vat
            map_transaction_out[product_code].including_vat[0] += including_vat

    # print('map_transaction_out', map_transaction_out)
    wb.close()
    return map_transaction_out

def read_transaction_in(map_master: dict, date: datetime.datetime, input_folder: str = Path.FOLDER_INPUT):
    """
    read transaction excel in Path.FOLDER_INPUT

    params: 
        map_master: Dict[str, Master] 
        input_folder: str = path of folder

    return: map_transaction_in: Dict[str, TransactionIn]
    """

    map_transaction_in: Dict[str, TransactionIn] = {}

    file_trx_in = os.path.join(input_folder, f"transaction in - {date.strftime('%Y %m')}.xlsx")

    check_file_exist(file=file_trx_in)

    # open workbook
    wb = load_workbook(filename=file_trx_in)
    ws = wb[wb.sheetnames[0]]

    # read header
    header = get_header_by_row(ws=ws, row_index=1)

    for row in ws.iter_rows(min_row=2):
        row = readrow(row)

        product_code = row[header.index(TransactionOutField.PRODUCT_CODE)]
        product_name = row[header.index(TransactionOutField.PRODUCT_NAME)]
        unit = row[header.index(TransactionOutField.UNIT)]
        qty = row[header.index(TransactionOutField.QTY)]

        if product_code not in map_master:
            continue

        if product_code not in map_transaction_in:
            map_transaction_in[product_code] = TransactionIn(
                product_code=product_code,
                product_name=product_name,
                unit=unit,
                qty=qty,
            )

        else:
            map_transaction_in[product_code].qty[0] += qty

    wb.close()
    return map_transaction_in

def write_yearly_report(map_transaction_out: Dict[str, TransactionOut], map_master: Dict[str, Master], date:datetime.datetime):

    wb = Workbook()
    ws = wb.active
    # ws.title = "Jan 2022"
    ws.title = date.strftime('%b %Y')

    # set header
    ws['A2'] = f'รายงานสรุปรายการขายสินค้า ประจำเดือน {get_full_month_th_with_year(month=date.month, year=date.year)}'

    headers = TransactionOutField.MAP_HEADERS

    map_summary = {
        TransactionOutField.VAT: 0,
        TransactionOutField.EXCLUDING_VAT: 0,
        TransactionOutField.INCLUDING_VAT: 0,
    }

    for at_column, header in enumerate(headers, start=1):
        ws.cell(row=4, column=at_column).value = header

    sorted_product_code = sorted(list(map_master.keys()))
    last_row = 4
    # set data
    for at_row, product_code in enumerate(sorted_product_code, start=5):
        # get TransactionOut
        last_row += 1
        trx = map_transaction_out[product_code]
        for at_column, header_mapper in enumerate(headers.values(), start=1):
            value = map_header_value(object_attributes=trx.__dict__, header_mapper=header_mapper)

            ws.cell(row=at_row, column=at_column).value = value
            

            # increase value
            if header_mapper in map_summary:
                map_summary[header_mapper] += value


    # set summary
    summary_row = last_row + 2
    
    ws.cell(row=summary_row, column=2).value = 'รวม'
    ws.cell(row=summary_row, column=5).value = f'=SUM(E5:E{last_row})'
    ws.cell(row=summary_row, column=6).value = f'=SUM(F5:F{last_row})'
    ws.cell(row=summary_row, column=7).value = f'=SUM(G5:G{last_row})'


    # save
    save_file=os.path.join(Path.FOLDER_OUTPUT, f'รายงานสรุปรายการขายสินค้าแต่ละเดือน ปี 2022.xlsx')
    overwrite_excel(wb=wb, save_file=save_file )

    wb.close()


align_center = Alignment(horizontal='center', vertical='center')

class InventYearlyReport:

    def __init__(self, save_file:str) -> None:
        self.save_file = save_file
        self.map_month_headers = {}
        self.map_summary_headers = {}
        self.map_inventory = {}
        self.map_summary = {}
        

    def initial_workbook(self):
        self.wb = Workbook()

    def save(self):
        overwrite_excel(wb=self.wb, save_file=self.save_file )
        self.wb.close()

    def __set_header_and_merge_below(self, headers: list, at_row: int = 4, start_col:int = 1, map_column: dict = {}):
        for at_col, header in enumerate(headers):
            self.ws.cell(row=at_row, column=start_col + at_col).value = header
            self.ws.merge_cells(start_row=at_row, start_column=start_col + at_col, end_row=at_row + 1, end_column=start_col + at_col)
            self.ws.cell(row=at_row, column=start_col + at_col).alignment = align_center
            map_column[header] = start_col + at_col

    def set_constant_cells(self, year:int):
        self.ws = self.wb.active
        self.ws.title = 'Summary'
        self.ws['A2'] = f"รายงานสรุปจำนวนสินค้าคงคลัง ประจำปี {year}"
        master_headers = ('รหัสสินค้า', 'ชื่อสินค้า', 'ราคาต่อหน่วย', f'จำนวนสินค้ายกยอดจากปี {year}')
        
        # Write master header
        self.__set_header_and_merge_below(headers=master_headers, at_row=4, start_col=1)

        # write monthly header
        start_month_col = len(master_headers)  
        month_column = start_month_col
        sub_month_headers = ("ยกยอดมา",	"เบิกออก",	"รับเข้า",	"คงเหลือ")

        for month in range(1, 13, 1):
            month_eng_short_name = datetime.datetime(year=year, month=month ,day=1).strftime('%b')

            self.map_month_headers[month] = {}
            start_column = month_column + 1
            end_column = month_column + 4
            self.ws.cell(row=4, column=start_column ).value = month_eng_short_name
            self.ws.merge_cells(start_row=4, start_column=start_column, end_row=4, end_column=end_column)
            self.ws.cell(row=4, column=start_column ).alignment = align_center
            
            for i, sub_header in enumerate(sub_month_headers):
                self.ws.cell(row=5, column=start_column + i).value = sub_header
                self.ws.cell(row=5, column=start_column + i).alignment = align_center

                self.map_month_headers[month][sub_header] = start_column + i

            month_column = end_column

        sum_headers = ('รวมรับเข้า', 'รวมเบิกออก', 'รวมคงเหลือ', 'มูลค่าสินค้าคงคลัง')
        self.__set_header_and_merge_below(headers=sum_headers, at_row=4, start_col=month_column + 2, map_column=self.map_summary_headers)

    def set_data(self, 
                 map_master: Dict[str, Master], 
                 current_map_transaction_out:Dict[str, TransactionOut], 
                 current_map_transaction_in: Dict[str, TransactionIn], 
                 month: int, 
                 start_row: int = 6,
                 ):
        
        self.map_inventory[month] = {}

        for at_row, product_code in enumerate(map_master, start=start_row):
            master = map_master[product_code]
            trx_out = current_map_transaction_out.get(product_code)
            trx_in = current_map_transaction_in.get(product_code)

            # set master
            self.ws.cell(row=at_row, column=1).value = product_code
            self.ws.cell(row=at_row, column=2).value = master.product_name
            self.ws.cell(row=at_row, column=3).value = master.price_per_unit
            self.ws.cell(row=at_row, column=4).value = 0

            if month == 1:
                last_stock = 0
                qty_in = get_qty_from_transaction(trx_in) 
                qty_out = get_qty_from_transaction(trx_out)
                balance = qty_in + last_stock - qty_out  
            else:
                last_stock = self.map_inventory[month-1][product_code]["คงเหลือ"]
                qty_in = get_qty_from_transaction(trx_in) 
                qty_out = get_qty_from_transaction(trx_out)
                balance = qty_in + last_stock - qty_out 

            self.map_inventory[month][product_code] = {
                "ยกยอดมา": last_stock,
                "เบิกออก": qty_out,	
                "รับเข้า": qty_in,	
                "คงเหลือ": balance
            }


            if product_code not in self.map_summary:
                self.map_summary[product_code] = {
                        "เบิกออก": qty_out,	
                        "รับเข้า": qty_in,	
                    }
            else:
                self.map_summary[product_code]["เบิกออก"] += qty_out
                self.map_summary[product_code]["รับเข้า"] += qty_in
            

            self.ws.cell(row=at_row, column=self.map_month_headers[month]['ยกยอดมา']).value = last_stock
            self.ws.cell(row=at_row, column=self.map_month_headers[month]['เบิกออก']).value = qty_out
            self.ws.cell(row=at_row, column=self.map_month_headers[month]['รับเข้า']).value = qty_in
            self.ws.cell(row=at_row, column=self.map_month_headers[month]['คงเหลือ']).value = balance 



    def calculate_summary(self, map_master: Dict[str, Master],  start_row: int = 6 ):

        total_stock_value = 0
        last_row = start_row
        for at_row, product_code in enumerate(map_master, start=start_row):
            last_row = at_row
            master = map_master[product_code]

            total_in =  self.map_summary[product_code]["รับเข้า"]
            total_out = self.map_summary[product_code]["เบิกออก"]
            stock_value = self.map_inventory[12][product_code]['คงเหลือ'] * master.price_per_unit

            self.ws.cell(row=at_row, column=self.map_summary_headers['รวมรับเข้า']).value = total_in
            self.ws.cell(row=at_row, column=self.map_summary_headers['รวมเบิกออก']).value = total_out
            self.ws.cell(row=at_row, column=self.map_summary_headers['รวมคงเหลือ']).value = self.map_inventory[12][product_code]['คงเหลือ']
            self.ws.cell(row=at_row, column=self.map_summary_headers['มูลค่าสินค้าคงคลัง']).value = stock_value
            total_stock_value += stock_value


        self.ws.cell(row=last_row + 1 , column=self.map_summary_headers['มูลค่าสินค้าคงคลัง']).value = total_stock_value
        self.ws.cell(row=last_row + 1 , column=self.map_summary_headers['มูลค่าสินค้าคงคลัง']).number_format = '#,##0.00'


            


def get_qty_from_transaction(obj):
    return 0 if obj is None else obj.qty[0]

def test(trx_in_qty):
    print(trx_in_qty)
    return trx_in_qty

def main():

    # Read master
    map_master = read_master()

    # Read transaction out per month
    save_file=os.path.join(Path.FOLDER_OUTPUT, f'รายงานสรุปรายการขายสินค้าแต่ละเดือน ปี 2022.xlsx')
    
    yearly_report = YearlyReport(save_file=save_file)
    yearly_report.initial_workbook()

    for month in range(1,13):

        date = datetime.datetime(2022,month, 1)
        print('date', date)

        map_transaction_out = read_transaction_out(map_master=map_master, date=date)
        
        yearly_report.set_constant_cells(date=date)
        yearly_report.set_data(map_master=map_master, map_transaction_out=map_transaction_out)
        yearly_report.set_summary()

    yearly_report.remove_sheet()
    yearly_report.save()


    # inventory entire year

    year = 2022
    save_file = os.path.join(Path.FOLDER_OUTPUT, f'รายงานสรุปจำนวนสินค้าคงคลังทั้งปี {year}.xlsx')

    inventory_yearly_report = InventYearlyReport(save_file=save_file)
    inventory_yearly_report.initial_workbook()
    inventory_yearly_report.set_constant_cells(year=year)


    for month in range(1,13):

        date = datetime.datetime(year, month, 1)
        print('date', date)
        
        current_map_transaction_out = read_transaction_out(map_master=map_master, date=date) # now 2
        current_map_transaction_in = read_transaction_in(map_master=map_master, date=date)

        inventory_yearly_report.set_data(
            map_master=map_master, 
            current_map_transaction_out=current_map_transaction_out, 
            current_map_transaction_in=current_map_transaction_in,
            month=date.month, 
            start_row= 6
            )

    inventory_yearly_report.calculate_summary(map_master=map_master)
    inventory_yearly_report.save()

if __name__ == '__main__':

    print('running')
    main()
    print('end')
    
    # for month in TH_MONTHS:
    #     print(month)
    # pass
