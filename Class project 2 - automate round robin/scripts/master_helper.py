from config import Path

from openpyxl import load_workbook

from excel_utils import readrow


class Product:
  product_code = ''
  product_name = ''
  group = ''
  min_tier = ''

  def __init__(self, product_code, product_name, group, min_tier: int) -> None:
    self.product_code = product_code
    self.product_name = product_name
    self.group = group
    self.min_tier = min_tier

  def __repr__(self) -> str:
    return f"{self.product_name} {self.group} tier:{self.min_tier}"
  





def read_master(file):

  # ====================================== sub function =======================================
  def get_product_detail(ws):
    map_product = {}

    for row in ws.iter_rows(min_row=2):
      row = readrow(row)
      
      product_code = row[0]
      product_name = row[1]
      group = row[2]
      min_tier = row[3]
      map_product[product_code] = Product(product_code, product_name, group, min_tier)

    return map_product
  
  def get_income_detail(ws):
    map_income_range = {}

    for row in ws.iter_rows(min_row=2):
      row = readrow(row)
      
      tier = row[0]
      income_range = row[1]
      map_income_range[income_range] = tier
    
    return map_income_range

  def get_sale_member_detail(ws):
    map_sale_group = {}

    for row in ws.iter_rows(min_row=2):
      row = readrow(row)
      
      sale_id = row[0]
      sale_name = row[1]
      slae_group = row[2]
      if slae_group not in map_sale_group:
        map_sale_group[slae_group] = []
      map_sale_group[slae_group].append((sale_id, sale_name))
    
    return map_sale_group

  # ====================================== process =======================================
  wb = load_workbook(file)

   

  # read product
  map_product = get_product_detail(ws=wb['Product code'])

  # read income range
  map_income_range = get_income_detail(ws=wb['Income range'])
  
  # read sale member
  map_sale_group = get_sale_member_detail(ws=wb['Sale member'])

  
  wb.close()

  return map_product, map_income_range, map_sale_group
 
# def read_master(file):

#   wb = load_workbook(file)

#   map_product = {}
#   map_income_range = {}
#   map_sale_group = {}

#   # read product
#   ws = wb['Product code']
#   map_product = get_product_detail(ws=ws)

#   for row in ws.iter_rows(min_row=2):
#     row = readrow(row)
    
#     product_code = row[0]
#     product_name = row[1]
#     group = row[2]
#     min_tier = row[3]
#     map_product[product_code] = Product(product_code, product_name, group, min_tier)


#   # read income range
#   ws = wb['Income range']
#   for row in ws.iter_rows(min_row=2):
#     row = readrow(row)
    
#     tier = row[0]
#     income_range = row[1]
#     map_income_range[income_range] = tier
  
  
#   # read sale member
#   ws = wb['Sale member']
#   for row in ws.iter_rows(min_row=2):
#     row = readrow(row)
    
#     sale_id = row[0]
#     sale_name = row[1]
#     slae_group = row[2]
#     if slae_group not in map_sale_group:
#       map_sale_group[slae_group] = []
#     map_sale_group[slae_group].append((sale_id, sale_name))

#   wb.close()

#   return map_product, map_income_range, map_sale_group


if __name__ == "__main__":
  
  map_product, map_income_range, map_sale_group = read_master(file=Path.MASTER_FILE)

  for k, v in map_product.items():
    print(k, v)

  print()

  for k, v in map_income_range.items():
    print(k, v)

  print()

  for k, v in map_sale_group.items():
    print(k, v)