
from openpyxl import load_workbook, Workbook
from config import Path
import os
from excel_utils import readrow, overwrite_excel
from typing import List

from master_helper import read_master, Product

class Registeration:
    def __init__(self, product: Product, customer_code, customer_name, customer_tel, income_range, register_date, customer_tier, file) -> None:
        self.product = product
        self.customer_code = customer_code
        self.customer_name = customer_name
        self.customer_tel = customer_tel
        self.income_range = income_range
        self.register_date = register_date
        self.customer_tier = customer_tier
        self.file = file
    
    def __repr__(self) -> str:
        return f"Registeration object code: {self.customer_code}"


def validate_income(product_tier: int, customer_tier: int):
    # validate data 
    # return 
    #   True = valid, 
    #   False = invalid
    return customer_tier >= product_tier



def read_input(file:str, map_product:dict, map_income_range:dict, map_valid_register_group:dict):
    wb = load_workbook(file)

    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2):
        row = readrow(row)

        product_code = row[1]
        product_name = row[2]	
        customer_code = row[3]	
        customer_name = row[4]	
        customer_tel = row[5]	
        income_range = row[6]	
        register_date = row[7]

        # validate data
        product: Product = map_product[product_code]
        customer_tier = map_income_range[income_range]
        # if customer_tier < product.min_tier:
        #     continue    # invalid data
        if not validate_income(product_tier=product.min_tier, customer_tier=customer_tier):
            continue
        
        # valid data
        # get group
        # print(customer_code,product_code, product.group)

        if product.group not in map_valid_register_group:
            map_valid_register_group[product.group] = []
        
        map_valid_register_group[product.group].append(Registeration(
              product=product, 
              customer_code=customer_code, 
              customer_name=customer_name, 
              customer_tel=customer_tel, 
              income_range=income_range, 
              register_date=register_date, 
              customer_tier=customer_tier,
              file=file
            ))

    wb.close()
    return map_valid_register_group





def round_robin(sales: list, valid_customer: List[Registeration]):
    copied_sales = sales.copy()

    output = []
    for data in valid_customer:
        data: Registeration

        if len(sales) == 0:
            sales = copied_sales.copy()

        # sales[0] is tuple of sale_id, sale_name
        x = [sales[0], data]
        output.append(x) 

        if len(sales) > 0:
            sales.pop(0)

    return output

def queue_valid_customer(map_valid_register_group: dict, map_sale_group: dict):
    map_all_output = {}

    for group in map_valid_register_group:

        if group not in map_all_output:
            map_all_output[group] = []

        map_all_output[group] = round_robin(sales=map_sale_group[group].copy() , valid_customer=map_valid_register_group[group])   

    return map_all_output


def write_output(file: str, map_all_output: dict):

    wb = Workbook()
    ws = wb.active
    headers = ['#', 'Product code', 'Product name'	, 'customer_code', 'name', 'tel', 'income_range', 'register_date', 'Sale ID'	, 'Sale Name', 'Group', 'from file']
    
    print('headers', headers)

    for group, data in map_all_output.items():

        sheet =  f"Group {group}"
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
        else:
            ws = wb[sheet]
        
        # set header
        ws.append(headers)
        for at_row, output in enumerate(data, start=2):
            output: List[tuple, Registeration]

            sale_id, sale_name = output[0]
            regis: Registeration = output[1]

            ws.cell(row=at_row, column=1).value = at_row - 1
            ws.cell(row=at_row, column=2).value = regis.product.product_code
            ws.cell(row=at_row, column=3).value = regis.product.product_name
            ws.cell(row=at_row, column=4).value = regis.customer_code
            ws.cell(row=at_row, column=5).value = regis.customer_name
            ws.cell(row=at_row, column=6).value = regis.customer_tel
            ws.cell(row=at_row, column=7).value = regis.income_range
            ws.cell(row=at_row, column=8).value = regis.register_date
            ws.cell(row=at_row, column=9).value = sale_id
            ws.cell(row=at_row, column=10).value = sale_name
            ws.cell(row=at_row, column=11).value = group
            ws.cell(row=at_row, column=12).value = regis.file.split("\\")[-1]
    

    wb.remove(wb['Sheet'])
    overwrite_excel(wb=wb, save_file=file)



if __name__ == "__main__":

    #  read master file
    map_product, map_income_range, map_sale_group = read_master(file=Path.MASTER_FILE)


    manaul_file = os.path.join(Path.FOLDER_INPUT, 'Data from manual.xlsx')

    file_inputs = [
        os.path.join(Path.FOLDER_INPUT, 'Data from manual.xlsx'),
        os.path.join(Path.FOLDER_INPUT, 'Data from Phone.xlsx'),
        os.path.join(Path.FOLDER_INPUT, 'Data from website.xlsx'),
    ]

    map_valid_register_group = {}
    for file in file_inputs:
        map_valid_register_group = read_input( 
            file=file, 
            map_product=map_product, 
            map_income_range=map_income_range, 
            map_valid_register_group=map_valid_register_group
            )

    map_all_output = queue_valid_customer(map_valid_register_group, map_sale_group)

    write_output(file=os.path.join(Path.FOLDER_OUTPUT, 'out.xlsx'), map_all_output=map_all_output)
    
    pass
